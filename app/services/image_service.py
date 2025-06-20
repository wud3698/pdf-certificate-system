import os
import base64
import zipfile
from io import BytesIO
from PIL import Image
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader

class ImageService:
    def __init__(self, app):
        self.app = app
        self.allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
        
    def allowed_file(self, filename):
        """检查文件扩展名是否允许"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in self.allowed_extensions
    
    def extract_images_from_excel(self, excel_file_path, image_column_index=None):
        """从Excel文件中提取图片"""
        try:
            # 确保图片存储目录存在
            image_folder = os.path.join(self.app.config['UPLOAD_FOLDER'], 'participant_images')
            os.makedirs(image_folder, exist_ok=True)
            
            # 打开Excel工作簿
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            
            # 创建图片加载器
            image_loader = SheetImageLoader(sheet)
            
            # 存储提取的图片信息
            image_data = {}
            
            # 如果指定了图片列索引，从该列提取图片
            if image_column_index is not None:
                column_letter = openpyxl.utils.get_column_letter(image_column_index + 1)
                
                for row in range(2, sheet.max_row + 1):  # 从第2行开始（跳过标题行）
                    cell = f"{column_letter}{row}"
                    dataframe_index = row - 2  # 正确的DataFrame索引映射
                    
                    try:
                        if image_loader.image_in(cell):
                            image = image_loader.get(cell)
                            
                            # 生成唯一的文件名
                            filename = f"participant_{dataframe_index}_{cell}.png"
                            image_path = os.path.join(image_folder, filename)
                            
                            # 保存图片
                            image.save(image_path, 'PNG')
                            
                            # 记录图片信息
                            image_data[dataframe_index] = {  # 使用正确的DataFrame索引
                                'filename': filename,
                                'path': image_path,
                                'relative_path': os.path.join('uploads', 'participant_images', filename)
                            }
                            
                            print(f"✅ 提取图片: Excel行{row} -> DataFrame索引{dataframe_index} -> {filename}")
                    except Exception as e:
                        print(f"⚠️  提取行{row}图片失败: {e}")
                        continue
            
            workbook.close()
            return image_data
            
        except Exception as e:
            print(f"❌ 从Excel提取图片失败: {e}")
            return {}
    
    def process_base64_image(self, base64_data, filename_prefix="image"):
        """处理base64编码的图片数据"""
        try:
            # 确保图片存储目录存在
            image_folder = os.path.join(self.app.config['UPLOAD_FOLDER'], 'participant_images')
            os.makedirs(image_folder, exist_ok=True)
            
            # 解码base64数据
            if base64_data.startswith('data:image'):
                # 移除data:image前缀
                base64_data = base64_data.split(',')[1]
            
            image_data = base64.b64decode(base64_data)
            
            # 使用PIL打开图片
            image = Image.open(BytesIO(image_data))
            
            # 生成文件名
            filename = f"{filename_prefix}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.png"
            image_path = os.path.join(image_folder, filename)
            
            # 保存图片
            image.save(image_path, 'PNG')
            
            return {
                'filename': filename,
                'path': image_path,
                'relative_path': os.path.join('uploads', 'participant_images', filename)
            }
            
        except Exception as e:
            print(f"❌ 处理base64图片失败: {e}")
            return None
    
    def resize_image(self, image_path, max_width=300, max_height=400):
        """调整图片大小"""
        try:
            with Image.open(image_path) as img:
                # 计算新的尺寸，保持宽高比
                img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                
                # 保存调整后的图片
                img.save(image_path, optimize=True, quality=85)
                
                return True
        except Exception as e:
            print(f"❌ 调整图片大小失败: {e}")
            return False
    
    def validate_image(self, image_path):
        """验证图片文件"""
        try:
            with Image.open(image_path) as img:
                # 检查图片格式
                if img.format.lower() not in ['png', 'jpeg', 'jpg', 'gif', 'bmp']:
                    return False, "不支持的图片格式"
                
                # 检查图片尺寸
                width, height = img.size
                if width > 2000 or height > 2000:
                    return False, "图片尺寸过大"
                
                return True, "图片验证通过"
                
        except Exception as e:
            return False, f"图片验证失败: {e}"
    
    def get_image_info(self, image_path):
        """获取图片信息"""
        try:
            with Image.open(image_path) as img:
                return {
                    'format': img.format,
                    'size': img.size,
                    'mode': img.mode,
                    'file_size': os.path.getsize(image_path)
                }
        except Exception as e:
            return None
    
    def delete_image(self, image_path):
        """删除图片文件"""
        try:
            if os.path.exists(image_path):
                os.remove(image_path)
                return True
        except Exception as e:
            print(f"❌ 删除图片失败: {e}")
        return False
    
    def save_uploaded_image(self, file, filename_prefix="image"):
        """保存上传的图片文件"""
        try:
            # 确保图片存储目录存在
            image_folder = os.path.join(self.app.config['UPLOAD_FOLDER'], 'participant_images')
            os.makedirs(image_folder, exist_ok=True)
            
            # 获取文件扩展名
            filename = secure_filename(file.filename)
            file_ext = os.path.splitext(filename)[1].lower()
            
            # 生成新的文件名
            timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            new_filename = f"{filename_prefix}_{timestamp}{file_ext}"
            
            # 构建文件路径
            file_path = os.path.join(image_folder, new_filename)
            
            # 保存文件
            file.save(file_path)
            
            # 验证保存的图片
            is_valid, message = self.validate_image(file_path)
            if not is_valid:
                # 如果图片无效，删除文件
                if os.path.exists(file_path):
                    os.remove(file_path)
                raise Exception(f"图片验证失败: {message}")
            
            # 调整图片大小（可选）
            self.resize_image(file_path, max_width=800, max_height=1000)
            
            return {
                'filename': new_filename,
                'path': file_path,
                'relative_path': os.path.join('uploads', 'participant_images', new_filename)
            }
            
        except Exception as e:
            print(f"❌ 保存上传图片失败: {e}")
            return None 